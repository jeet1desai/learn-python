# Install - pip install openpyxl
import openpyxl

inv_file = openpyxl.load_workbook("automation_sheet.xlsx")
product_list = inv_file["Product"]

# # How many row
# print(product_list.max_row)

# Ex:1 How many product per supplier
product_per_supplier = {}

for item_product in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(item_product, 4).value

    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] = product_per_supplier[supplier_name] + 1
    else:
        product_per_supplier[supplier_name] = 1

print(product_per_supplier)


# Ex:2 Total value inventory per supplier
inventory_per_supplier = {}

for item_product in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(item_product, 4).value
    inventory_value = product_list.cell(item_product, 2).value
    price = product_list.cell(item_product, 3).value

    if supplier_name in inventory_per_supplier:
        inventory_per_supplier[supplier_name] = inventory_per_supplier[supplier_name] + (price * inventory_value)
    else:
        inventory_per_supplier[supplier_name] = price * inventory_value

print(inventory_per_supplier)

# Ex: 3 Print product that have inventory less then 50
product_inventory_under_50 = {}

for item_product in range(2, product_list.max_row + 1):
    inventory_value = product_list.cell(item_product, 2).value
    product_num = product_list.cell(item_product, 1).value

    if inventory_value < 50:
        product_inventory_under_50[product_num] = inventory_value

print(product_inventory_under_50)

# Ex: 4 Add value in sheet
for item_product in range(2, product_list.max_row + 1):
    inventory_value = product_list.cell(item_product, 2).value
    price_value = product_list.cell(item_product, 3).value
    inventory_price = product_list.cell(item_product, 5)

    inventory_price.value = price_value * inventory_value

inv_file.save("automation_sheet.xlsx")



